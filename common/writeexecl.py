from openpyxl import load_workbook
import openpyxl


def copy_excel(e1, e2):
    wb2 = openpyxl.Workbook()
    wb2.save(e2)
    # 读取数据
    wb1 = openpyxl.load_workbook(e1)
    wb2 = openpyxl.load_workbook(e2)
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    sheet1 = wb1[sheets1[0]]
    sheet2 = wb2[sheets2[0]]
    max_row = sheet1.max_row            # 最大行数
    max_column = sheet1.max_column      # 最大列数
    for m in list(range(1, max_row+1)):
        for n in list(range(65, 65+max_column)):   # chr(97) = 'a'
            n = chr(n)                             # ASCII字符
            i = ('%s%d' % (n, m))
            cell1 = sheet1[i].value
            sheet2[i].value = cell1
    wb2.save(e2)
    wb1.close()
    wb2.close()


class WriteExcel(object):
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active

    def write(self,row_n, col_n, value):
        """写入数据，如(2,3，"hello"),第二行第三列写入数据"hello"""
        self.ws.cell(row_n, col_n).value = value
        self.wb.save(self.filename)


if __name__ == "__main__":
    copy_excel("testdata.xlsx", "debug.xlsx")
    wt = WriteExcel("debug.xlsx")
    wt.write(2, 1, "张三007")